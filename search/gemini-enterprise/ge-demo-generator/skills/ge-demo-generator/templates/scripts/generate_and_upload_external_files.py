#!/usr/bin/env python3
# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Deployed as a runtime template into the user's Cloud Shell (not imported by
# repo tooling); validated by py_compile and end-to-end demo deployments.
# Repo-level strict lint/typing is intentionally skipped for this generated-
# origin runtime code; incremental typing is planned as follow-up.
# flake8: noqa
# pylint: skip-file
# mypy: ignore-errors
# ruff: noqa


"""
Generates external demo files (PDF Audit Report, Excel Spreadsheet Ledger, Simulated Scanned Images)
and uploads them to the Google Drive of the deploy target - the active gcloud account.

The upload talks to the Drive v3 REST API with the access token `gcloud` already
holds, so there is no CLI to install and no second identity: the folder is owned
by the same account that owns the demo, which removes the cross-domain share that
used to fail more often than it worked. The one prerequisite is the Drive scope,
which a plain `gcloud auth login` does not grant - see get_drive_identity().
"""

import os
import sys
import json
import mimetypes
import subprocess
import argparse
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

DRIVE_API = "https://www.googleapis.com/drive/v3"
DRIVE_UPLOAD_API = "https://www.googleapis.com/upload/drive/v3"
FOLDER_MIME = "application/vnd.google-apps.folder"
REAUTH_HINT = "gcloud auth login --enable-gdrive-access"

def run_cmd(cmd, check=False):
    """Run a shell command and return stdout."""
    res = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if check and res.returncode != 0:
        print(f"Error running command: {cmd}\nStderr: {res.stderr}", file=sys.stderr)
    return res.stdout.strip(), res.returncode

def generate_pdf(output_path: str, title: str, sections: list, discrepancy_info: dict):
    """Generates a professional multi-section PDF document using reportlab with Japanese CID font support."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        # Register Japanese CID fonts to guarantee clean CJK rendering without tofu (■)
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))

        doc = SimpleDocTemplate(output_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle',
            fontName='HeiseiKakuGo-W5',
            fontSize=16,
            leading=20,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=15
        )
        h2_style = ParagraphStyle(
            'DocH2',
            fontName='HeiseiKakuGo-W5',
            fontSize=12,
            leading=16,
            textColor=colors.HexColor('#0369a1'),
            spaceBefore=12,
            spaceAfter=6
        )
        body_style = ParagraphStyle(
            'DocBody',
            fontName='HeiseiKakuGo-W5',
            fontSize=9.5,
            leading=14,
            textColor=colors.HexColor('#334155'),
            spaceAfter=8
        )

        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 10))

        for sec in sections:
            elements.append(Paragraph(sec.get('heading', ''), h2_style))
            elements.append(Paragraph(sec.get('content', ''), body_style))
            elements.append(Spacer(1, 6))

        if discrepancy_info and 'table_data' in discrepancy_info:
            elements.append(Paragraph(
                discrepancy_info.get('heading', 'Audit Discrepancy Summary'), h2_style))
            table_data = discrepancy_info['table_data']
            t = Table(table_data, colWidths=[110, 130, 110, 130])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'HeiseiKakuGo-W5'),
                ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 12))

        doc.build(elements)
        print(f"  ✅ Generated PDF: {output_path}")
        return True
    except Exception as e:
        print(f"  ⚠️ PDF generation failed: {e}", file=sys.stderr)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"# {title}\n\n")
            for sec in sections:
                f.write(f"## {sec.get('heading', '')}\n{sec.get('content', '')}\n\n")
        return False

# Wording used to dress the generated scan. The defaults are deliberately
# industry- and language-neutral: GE Demo Generator builds demos for any domain,
# so nothing here may assume logistics, manufacturing, retail or any other
# vertical. Pass a `style` dict (see --spec-file) to match the demo's domain,
# country and conversation language.
DEFAULT_DOC_STYLE = {
    "handwriting_language": "the same language as the printed header",
    "routine_doc_kind": "routine operational record form",
    "exception_doc_kind": "exception review form",
    "location": "",
    "approval_seal": "APPROVED",
    "review_stamp": "NEEDS REVIEW / PENDING APPROVAL",
    "urgent_note": "Does not match the system of record - urgent re-check requested.",
    # Labels for the Pillow fallback rendering (used when image generation fails).
    "label_issuer": "Issued by",
    "label_doc_no": "Ref No",
    "label_date": "Date",
    "label_review_note": "Review note",
    "label_disclaimer": "This document was generated automatically for a Gemini Enterprise demo ({company}). It is not a real record.",
}


def generate_document_image(output_path: str, title: str, company_name: str, doc_no: str, date_str: str, headers: list, rows: list, is_discrepancy: bool = False, project_id: str = None, style: dict = None):
    """Generates a photorealistic top-down scanned handwritten document image using Gemini 3.1 Flash Image, with fallback to Pillow."""
    st = dict(DEFAULT_DOC_STYLE)
    st.update(style or {})
    location = st["location"] or company_name

    # Attempt 1: Vertex AI Agent Platform image generation (gemini-3.1-flash-image)
    try:
        from google import genai
        from google.genai import types

        target_project = project_id or os.environ.get("PROJECT_ID") or os.environ.get("GOOGLE_CLOUD_PROJECT")
        if not target_project:
            raise RuntimeError("PROJECT_ID / GOOGLE_CLOUD_PROJECT is not set")
        client = genai.Client(vertexai=True, project=target_project, location="global")

        row_lines = []
        for r_i, r in enumerate(rows):
            row_lines.append(f"{r_i + 1}) " + " | ".join(str(v) for v in r))
        rows_text = "\n".join(row_lines)
        cols_text = " | ".join(headers)

        if not is_discrepancy:
            prompt = f"""A highly detailed, realistic top-down flat-lay scan of an authentic paper {st['routine_doc_kind']} ({company_name}) on a textured, slightly wrinkled sheet of paper filling the entire frame with zero background.
At the top: bold printed formal header "{title}", Date: "{date_str}", Ref No: "{doc_no}", Company: "{company_name}".
In the center: a printed table grid with column headers: {cols_text}.
Inside the table cells, hurried and realistic human handwriting in dark blue ballpoint pen ink, written in {st['handwriting_language']}:
{rows_text}
At the bottom right: a designated verification box carrying a red ink corporate approval stamp reading "{st['approval_seal']}".
Authentic paper grain, slight pen pressure indentations, real-world operational document photograph, sharp focus, top-down perspective."""
        else:
            prompt = f"""A highly detailed, realistic top-down flat-lay scan of an authentic paper {st['exception_doc_kind']} ({company_name}) on textured paper filling the frame.
At the top: bold printed formal header "{title}", Ref No: "{doc_no}", Location: "{location}".
In the center: a printed table grid with columns: {cols_text}.
Inside the table cells, messy hurried human handwriting in blue and black ballpoint pen ink, written in {st['handwriting_language']}:
{rows_text}
In the margin: a highlighted red pen handwritten urgent inspection note "{st['urgent_note']}".
Red ink stamp "{st['review_stamp']}" stamped on the sheet. Realistic lighting, paper texture, authentic operational document photograph."""

        for model_name in ["gemini-3.1-flash-image", "gemini-3-pro-image"]:
            try:
                res = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=types.GenerateContentConfig(response_modalities=["IMAGE"])
                )
                if res.candidates and res.candidates[0].content and res.candidates[0].content.parts:
                    for part in res.candidates[0].content.parts:
                        if part.inline_data and part.inline_data.data:
                            with open(output_path, "wb") as f:
                                f.write(part.inline_data.data)
                            print(f"  ✅ Generated Photorealistic Handwritten Document Image ({model_name}): {output_path}")
                            return True
            except Exception as m_err:
                pass
    except Exception as e:
        print(f"  ⚠️ Gemini Image generation skipped ({e}), falling back to Pillow...", file=sys.stderr)

    # Attempt 2: Fallback to Pillow
    try:
        from PIL import Image, ImageDraw, ImageFont

        width, height = 1000, 1350
        img = Image.new("RGB", (width, height), color="#FAF9F6")
        draw = ImageDraw.Draw(img)

        cjk_candidates = [
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf"
        ]
        font_file = None
        for candidate in cjk_candidates:
            if os.path.exists(candidate):
                font_file = candidate
                break

        if font_file:
            f_title = ImageFont.truetype(font_file, 24)
            f_sub = ImageFont.truetype(font_file, 14)
            f_body = ImageFont.truetype(font_file, 15)
            f_stamp = ImageFont.truetype(font_file, 18)
            f_small = ImageFont.truetype(font_file, 12)
        else:
            f_title = f_sub = f_body = f_stamp = f_small = ImageFont.load_default()

        draw.rectangle([(40, 40), (960, 1310)], fill="#FFFFFF", outline="#94A3B8", width=2)
        draw.rectangle([(60, 60), (940, 140)], fill="#F8FAFC", outline="#CBD5E1", width=1)
        draw.text((80, 75), title, fill="#0F172A", font=f_title)
        draw.text((80, 110), f"{st['label_issuer']}: {company_name} | {st['label_doc_no']}: {doc_no} | {st['label_date']}: {date_str}", fill="#475569", font=f_sub)
        draw.ellipse([(830, 65), (915, 135)], outline="#DC2626", width=3)
        draw.text((845, 88), st['approval_seal'] if not is_discrepancy else st['review_stamp'], fill="#DC2626", font=f_stamp)

        y_start = 180
        row_height = 65
        col_x = [60, 110, 480, 660, 800, 940]

        draw.rectangle([(60, y_start), (940, y_start + 45)], fill="#E2E8F0", outline="#94A3B8", width=1)
        for i, h in enumerate(headers):
            draw.text((col_x[i] + 10, y_start + 12), h, fill="#0F172A", font=f_body)

        cur_y = y_start + 45
        for r_idx, r in enumerate(rows):
            fill_bg = "#FFFBEB" if (is_discrepancy and r_idx == 0) else ("#FFFFFF" if r_idx % 2 == 0 else "#F8FAFC")
            draw.rectangle([(60, cur_y), (940, cur_y + row_height)], fill=fill_bg, outline="#E2E8F0", width=1)
            for c_idx, val in enumerate(r):
                if c_idx < len(col_x) - 1:
                    text_color = "#DC2626" if (is_discrepancy and r_idx == 0 and c_idx >= 3) else "#1E293B"
                    draw.text((col_x[c_idx] + 10, cur_y + 20), str(val), fill=text_color, font=f_body)
            for x in col_x:
                draw.line([(x, cur_y), (x, cur_y + row_height)], fill="#E2E8F0", width=1)
            cur_y += row_height

        if is_discrepancy:
            draw.rectangle([(60, cur_y + 30), (940, cur_y + 120)], fill="#FEF2F2", outline="#EF4444", width=1)
            draw.text((80, cur_y + 45), st['label_review_note'], fill="#991B1B", font=f_body)
            draw.text((80, cur_y + 75), st['urgent_note'], fill="#B91C1C", font=f_small)

        draw.text((60, 1260), st['label_disclaimer'].format(company=company_name), fill="#94A3B8", font=f_small)
        img.save(output_path, "JPEG", quality=92)
        print(f"  ✅ Generated Fallback Document Image: {output_path}")
        return True
    except Exception as e:
        print(f"  ⚠️ Document Image generation failed: {e}", file=sys.stderr)
        return False

def generate_excel(output_path: str, title: str, kpis: list, headers: list, rows: list):
    """Generates a rich Excel workbook using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "External_Ledger"

        title_font = Font(name='Calibri', size=16, bold=True, color='0F172A')
        kpi_font = Font(name='Calibri', size=11, bold=True, color='0369A1')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
        border_side = Side(style='thin', color='CBD5E1')
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.cell(row=1, column=1, value=title).font = title_font
        ws.row_dimensions[1].height = 28

        row_idx = 3
        if kpis:
            ws.cell(row=row_idx, column=1, value="[Executive Summary KPIs]").font = Font(bold=True, size=10, color='64748B')
            row_idx += 1
            for i, kpi in enumerate(kpis, start=1):
                cell = ws.cell(row=row_idx, column=i, value=f"{kpi.get('label', '')}: {kpi.get('value', '')}")
                cell.font = kpi_font
                cell.fill = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            row_idx += 2

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

        for r in rows:
            for col_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
            row_idx += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        print(f"  ✅ Generated Excel Ledger: {output_path}")
        return True
    except Exception as e:
        print(f"  ⚠️ Excel generation failed: {e}", file=sys.stderr)
        return False

def get_active_account() -> str:
    """Retrieves the active Google Cloud / Google Drive target account email."""
    try:
        res = subprocess.run(["gcloud", "config", "get-value", "account"], capture_output=True, text=True)
        lines = [l.strip() for l in res.stdout.strip().split("\n") if l.strip() and not l.startswith("Your active configuration")]
        if lines:
            return lines[-1]
    except Exception:
        pass
    return os.environ.get("GCP_ACCOUNT") or os.environ.get("USER") or "current user"


def drive_access_token() -> str:
    """The gcloud access token, or '' when gcloud cannot mint one."""
    out, code = run_cmd("gcloud auth print-access-token")
    return out.strip() if code == 0 else ""


def drive_request(token: str, method: str, url: str, body=None, content_type="", raw=b""):
    """One Drive v3 call. Returns (parsed_json_or_{}, status, error_text).

    Errors are values, not exceptions: every caller has something useful to do
    with a failure - name it in the skip reason, record it as share_error, or
    carry on without a link - and none of them should take the document
    generation down with them.
    """
    data = raw if raw else (json.dumps(body).encode("utf-8") if body is not None else None)
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    if data is not None:
        req.add_header("Content-Type", content_type or "application/json; charset=UTF-8")
    try:
        with urllib.request.urlopen(req, timeout=120) as res:
            payload = res.read().decode("utf-8", "replace")
            return (json.loads(payload) if payload.strip() else {}), res.status, ""
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", "replace")
        try:
            message = json.loads(detail).get("error", {}).get("message", "")
        except Exception:
            message = ""
        return {}, e.code, (message or detail.strip().splitlines()[0] if detail.strip() else str(e))[:300]
    except Exception as e:  # DNS, proxy, timeout
        return {}, 0, str(e)[:300]


def get_drive_identity(token: str):
    """(email, skip_reason) for the Drive the token can write into.

    `about.get` is the whole pre-flight: it answers who the token is AND whether
    it carries the Drive scope, which is the failure everyone hits. A plain
    `gcloud auth login` grants cloud-platform and nothing else, so Drive returns
    403 ACCESS_TOKEN_SCOPE_INSUFFICIENT and the fix is a re-login with
    --enable-gdrive-access. Saying that here is the difference between a demo
    that is one command from having its documents in Drive and one whose
    operator concludes the feature is broken.
    """
    if not token:
        return "", ("no gcloud access token is available on this machine "
                    f"(run: {REAUTH_HINT})")
    info, status, err = drive_request(token, "GET", f"{DRIVE_API}/about?fields=user")
    if status == 200:
        email = (info.get("user") or {}).get("emailAddress", "").strip()
        if email:
            return email, ""
        return "", "the Drive API returned no user for these credentials"
    if status == 403 and "scope" in (err or "").lower():
        return "", ("the gcloud credentials have no Google Drive scope - "
                    f"run `{REAUTH_HINT}` and re-run this script")
    if status in (401, 403):
        return "", f"Google Drive refused these credentials ({status}: {err})"
    return "", f"the Drive API could not be reached ({status or 'no response'}: {err})"


def drive_escape_query(value: str) -> str:
    """Escape a literal for a Drive `q` string.

    Company names with an apostrophe would otherwise terminate the quoted term
    early, and the query stops meaning what it says.
    """
    return value.replace("\\", "\\\\").replace("'", "\\'")


def drive_find_folder(token: str, name: str) -> str:
    """The id of a non-trashed folder this account OWNS with this exact name, or ''.

    `'me' in owners` matters: without it a folder someone else happened to share
    under the same name wins the search, and the upload then writes the demo
    documents into a stranger's Drive.
    """
    q = (f"name = '{drive_escape_query(name)}' and mimeType = '{FOLDER_MIME}' "
         "and trashed = false and 'me' in owners")
    url = f"{DRIVE_API}/files?" + urllib.parse.urlencode(
        {"q": q, "fields": "files(id)", "pageSize": "1"})
    info, status, _ = drive_request(token, "GET", url)
    files = info.get("files") or []
    return files[0].get("id", "") if status == 200 and files else ""


def drive_find_child(token: str, name: str, parent_id: str) -> str:
    """The id of a non-trashed file already called `name` inside `parent_id`, or ''.

    The folder is found by name, so a second deploy of the same suffix - a heal,
    a retry, a re-run after editing the content - lands in the folder that is
    already there. Without this lookup it would collect a second copy of all
    four documents every time. Replacing the file instead also keeps its id, so
    a link already pasted into a demo script still resolves.
    """
    if not parent_id:
        return ""
    q = (f"name = '{drive_escape_query(name)}' and '{parent_id}' in parents "
         "and trashed = false")
    url = f"{DRIVE_API}/files?" + urllib.parse.urlencode(
        {"q": q, "fields": "files(id)", "pageSize": "1"})
    info, status, _ = drive_request(token, "GET", url)
    files = info.get("files") or []
    return files[0].get("id", "") if status == 200 and files else ""


def drive_upload_file(token: str, path: str, parent_id: str):
    """Upload one file into parent_id with a multipart/related request.

    Returns (file_id, error). The error travels back because a folder full of
    nothing is the one outcome that must not be reported as success, and only
    the caller can see that every file failed the same way.

    A file of the same name already in the folder is replaced rather than
    duplicated - see drive_find_child(). `parents` is only writable on create,
    so the update sends the name alone.

    The four sample documents are a few hundred KB in total, so a single
    multipart request is the right shape - a resumable session would be three
    round trips per file to protect against an interruption that costs one
    second to retry.
    """
    name = os.path.basename(path)
    existing_id = drive_find_child(token, name, parent_id)
    metadata = {"name": name}
    if parent_id and not existing_id:
        metadata["parents"] = [parent_id]
    mime = mimetypes.guess_type(path)[0] or "application/octet-stream"
    boundary = "ge-demo-external-files-boundary"
    with open(path, "rb") as fh:
        content = fh.read()
    body = b"".join([
        f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n".encode("utf-8"),
        json.dumps(metadata).encode("utf-8"),
        f"\r\n--{boundary}\r\nContent-Type: {mime}\r\n\r\n".encode("utf-8"),
        content,
        f"\r\n--{boundary}--\r\n".encode("utf-8"),
    ])
    if existing_id:
        method, url = "PATCH", f"{DRIVE_UPLOAD_API}/files/{existing_id}?uploadType=multipart&fields=id"
    else:
        method, url = "POST", f"{DRIVE_UPLOAD_API}/files?uploadType=multipart&fields=id"
    info, status, err = drive_request(
        token, method, url,
        content_type=f"multipart/related; boundary={boundary}", raw=body)
    if status == 200 and info.get("id"):
        return info["id"], ""
    print(f"    ⚠️ Upload failed ({status}): {err}", file=sys.stderr)
    return "", f"{status}: {err}"

def upload_to_google_drive(company_name: str, suffix: str, files_to_upload: list) -> dict:
    """Creates a Google Drive folder and uploads all external demo files into the deploy target's Drive."""
    target_account = get_active_account()
    folder_name = f"GE Demo - {company_name} ({suffix})"
    
    print("\n" + "=" * 80)
    print(f"👤 Target Deployment Account    : {target_account}")
    print(f"📁 Target Folder Name          : '{folder_name}'")
    print("=" * 80)
    
    folder_id = ""
    folder_url = ""
    shared_permissions = []
    share_error = ""

    # The upload runs as the deploy target itself.
    #
    # `gcloud` is already authenticated as the account that owns the demo, so
    # the folder is OWNED by that account and there is nothing to share with it.
    # Until v2.13.0 this used a separate CLI authenticated as the operator, which
    # meant every cross-account demo depended on a Drive share landing - and a
    # google.com -> customer-domain share is frequently refused by policy, so the
    # target got a link that 404s. Owning the folder outright removes that whole
    # class of failure along with the "Shared with me" caveats.
    #
    # SKIP_DRIVE_UPLOAD=1 stays out of Drive entirely; the documents then live in
    # ./external_files/ and the bucket, and the deploy banner says so.
    skip_drive = os.environ.get("SKIP_DRIVE_UPLOAD", "").strip().lower() in ("1", "true", "yes")
    token = "" if skip_drive else drive_access_token()
    drive_account, identity_error = ("", "SKIP_DRIVE_UPLOAD is set") if skip_drive \
        else get_drive_identity(token)

    should_upload = bool(drive_account)
    skip_reason = "" if should_upload else identity_error

    if skip_reason:
        print(f"ℹ️ Skipping the Google Drive upload: {skip_reason}.")
        print("  Demo files are kept in ./external_files/ and staged to GCS instead.")
        print("  Nothing after this step can create a Drive copy - it is a deploy-time step.")

    if should_upload:
        # Idempotent by name: a re-run of the deploy reuses the folder it made
        # the first time rather than leaving the operator with three of them.
        folder_id = drive_find_folder(token, folder_name)
        if folder_id:
            print(f"  ♻️ Reusing the existing Drive folder '{folder_name}'.")
        else:
            info, status, err = drive_request(
                token, "POST", f"{DRIVE_API}/files?fields=id",
                body={"name": folder_name, "mimeType": FOLDER_MIME})
            folder_id = info.get("id", "")
            if not folder_id:
                skip_reason = f"the Drive folder could not be created ({status}: {err})"
                should_upload = False
                print(f"  ⚠️ {skip_reason}")

        if folder_id:
            folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
            print(f"  ✅ Folder ready: {folder_url}")

            # Link sharing is a convenience for the demo audience, not a
            # requirement: the owner can always open it. Organizations routinely
            # forbid "anyone with the link", so a refusal is recorded and the
            # run continues - it is not the failed share of the old design,
            # where the target could not open its own demo documents.
            _, status, err = drive_request(
                token, "POST",
                f"{DRIVE_API}/files/{folder_id}/permissions?sendNotificationEmail=false",
                body={"type": "anyone", "role": "reader"})
            if status == 200:
                shared_permissions.append("Anyone with link (Reader)")
                print("  ✅ Link sharing enabled: Anyone with link (Reader)")
            else:
                share_error = f"{status}: {err}"
                print(f"  ℹ️ Link sharing not enabled ({share_error}). "
                      f"{drive_account} can still open the folder.")

    # owner_account is the account that actually OWNS the Drive resources. The
    # deploy banner prints it as "Drive Owner Account" and tells the operator
    # which Google account to be signed in as, so it must stay empty when
    # nothing was uploaded - otherwise the banner sends people to a folder that
    # does not exist. Since v2.13.0 the owner is the token's own identity, which
    # is normally the deploy target; they can still differ if `gcloud config
    # get-value account` is not the account that minted the token.
    results = {
        "owner_account": drive_account if folder_url else "",
        "target_account": target_account,
        "upload_skipped_reason": skip_reason,
        "shared_permissions": shared_permissions,
        "share_error": share_error,
        "folder_name": folder_name,
        "folder_id": folder_id,
        "folder_url": folder_url,
        "uploaded_files": []
    }

    uploaded = 0
    upload_error = ""
    for file_path in files_to_upload:
        if not os.path.exists(file_path):
            continue
        fname = os.path.basename(file_path)
        file_id = ""
        file_url = ""

        if folder_id and should_upload:
            print(f"  📤 Uploading '{fname}' to Google Drive...")
            file_id, upload_err = drive_upload_file(token, file_path, folder_id)
            if file_id:
                uploaded += 1
                file_url = f"https://drive.google.com/file/d/{file_id}/view"
                print(f"    ✅ Link: {file_url}")
            else:
                upload_error = upload_error or upload_err

        results["uploaded_files"].append({
            "fileName": fname,
            "fileId": file_id,
            "url": file_url or f"file://{os.path.abspath(file_path)}",
            "localPath": os.path.abspath(file_path)
        })

    # An empty folder is not a Drive copy. The folder create and the uploads are
    # separate permissions - a service-account token, for instance, may create
    # folders all day and have no storage quota to put anything in them - so
    # reporting the folder URL after every upload failed hands the demo a link
    # to nothing. Report it as no Drive copy, with the upload's own error, and
    # keep folder_id so cleanup.sh still trashes the empty folder.
    if should_upload and folder_id and uploaded == 0:
        skip_reason = f"the Drive folder was created but every upload failed ({upload_error})"
        results["upload_skipped_reason"] = skip_reason
        results["folder_url"] = ""
        results["owner_account"] = ""
        folder_url = ""
        print(f"  ⚠️ {skip_reason}")

    print("\n" + "-" * 80)
    print(f"👤 Target Account     : {target_account}")
    print(f"📂 Folder Link        : {folder_url or 'Stored Locally in ./external_files/'}")
    print(f"🔑 Permissions Granted: {', '.join(shared_permissions) if shared_permissions else 'Local file store'}")
    if folder_url:
        print(f"👑 Drive Owner        : {results['owner_account']}")
        print(f"⚠️ ACCESS INSTRUCTION : Switch browser to Google Account [{results['owner_account']}] before opening links.")
        if results["owner_account"].lower() != (target_account or "").lower():
            print(f"                        Note: the deploy target is {target_account}, a different account.")
            print(f"                        Share the folder with it from the Drive UI as {results['owner_account']}.")
        if share_error:
            print(f"ℹ️ LINK SHARING OFF   : 'anyone with the link' was refused ({share_error}).")
            print("                        The owner can open the folder; share it explicitly for others.")
    print("-" * 80 + "\n")

    return results

def main():
    parser = argparse.ArgumentParser(description="Generate and upload external sample demo files to Google Drive.")
    parser.add_argument("--domain", default="demo.example.com", help="Customer domain name")
    parser.add_argument("--company", default="Demo Company", help="Company Name")
    parser.add_argument("--suffix", default="1234", help="Unique demo suffix")
    parser.add_argument("--outdir", default="./external_files", help="Output directory for generated files")
    parser.add_argument(
        "--spec-file", default="",
        help="JSON file describing this demo's external files. Shape: "
             "{'style': {...}, 'pdf': {'title', 'sections', 'discrepancy'}, "
             "'excel': {'title', 'kpis', 'headers', 'rows'}, "
             "'scans': [{'title', 'doc_no', 'date', 'headers', 'rows', 'is_discrepancy'}]}. "
             "Everything is written in the demo's own language and domain; the built-in "
             "fallback is generic placeholder content.")
    args = parser.parse_args()

    out_dir = Path(args.outdir)
    out_dir.mkdir(parents=True, exist_ok=True)

    pdf_path = str(out_dir / f"{args.domain.replace('.', '_')}_audit_report.pdf")
    xlsx_path = str(out_dir / f"{args.domain.replace('.', '_')}_external_ledger.xlsx")
    img1_path = str(out_dir / f"{args.domain.replace('.', '_')}_simulated_order_task1.jpg")
    img2_path = str(out_dir / f"{args.domain.replace('.', '_')}_simulated_order_task2_discrepancy.jpg")

    # The demo's own content comes from --spec-file, written in Phase 3 by the
    # skill. GE Demo Generator targets ANY domain, so the built-in fallback
    # below is deliberately generic: no industry vocabulary, no locale, no
    # currency. Never replace it with content from one customer's demo.
    spec = {}
    if args.spec_file:
        if not os.path.exists(args.spec_file):
            print(f"❌ Spec file not found: {args.spec_file}", file=sys.stderr)
            return 1
        with open(args.spec_file, encoding="utf-8") as f:
            spec = json.load(f)

    style = spec.get("style", {})

    # 1. PDF report
    pdf_spec = spec.get("pdf", {})
    pdf_title = pdf_spec.get("title", f"{args.company} - External Reconciliation Report")
    pdf_sections = pdf_spec.get("sections", [
        {
            "heading": "1. Executive Summary",
            "content": "This report compares the records held in the internal system with the "
                       "documents received from external counterparties for the period under review."
        },
        {
            "heading": "2. Key Discrepancies",
            "content": "A small number of records differ between the two sources. The differences are "
                       "consistent with manual entry at the point of receipt or with a delay in the "
                       "downstream synchronization."
        },
        {
            "heading": "3. Recommendations",
            "content": "Cross-reference both sources automatically, update the records that can be "
                       "reconciled without judgement, and escalate the remainder for human approval."
        }
    ])
    discrepancy_info = pdf_spec.get("discrepancy", {
        "heading": "Discrepancy Summary",
        "table_data": [
            ["Reference ID", "Counterparty", "System Value", "Document Value"],
            ["REF-0001", "Counterparty A", "1,200", "1,050 (mismatch)"],
            ["REF-0002", "Counterparty B", "850", "850 (match)"],
            ["REF-0003", "Counterparty C", "2,400", "2,160 (mismatch)"],
            ["REF-0004", "Counterparty D", "500", "500 (match)"]
        ]
    })
    generate_pdf(pdf_path, pdf_title, pdf_sections, discrepancy_info)

    # 2. Excel ledger
    xl_spec = spec.get("excel", {})
    xlsx_title = xl_spec.get("title", f"{args.company} - External Records Ledger")
    kpis = xl_spec.get("kpis", [
        {"label": "Records Reviewed", "value": "50"},
        {"label": "Mismatch Rate", "value": "8.4%"},
        {"label": "Open Items", "value": "6"}
    ])
    headers = xl_spec.get("headers",
                          ["Reference_ID", "Counterparty", "Record_Date", "Amount",
                           "Status", "Notes"])
    rows = xl_spec.get("rows")
    if not rows:
        rows = []
        for i in range(1, 51):
            flagged = i in (1, 7, 14, 22, 35, 42)
            rows.append([
                f"REF-{100 + i:04d}",
                f"Counterparty {chr(65 + (i % 6))}",
                f"2026-03-{min(i % 28 + 1, 28):02d}",
                f"{150000 + i * 12500:,}",
                "FLAGGED_DISCREPANCY" if flagged else "CONFIRMED",
                "Differs from the internal record" if flagged else "Matches the internal record",
            ])
    generate_excel(xlsx_path, xlsx_title, kpis, headers, rows)

    # 3 & 4. Scanned forms - one routine, one carrying a deliberate discrepancy.
    scan_specs = spec.get("scans", [])
    default_headers = ["No.", "Description", "Reference", "Value", "Status"]
    scan_defaults = [
        {
            "title": f"{args.company} - Operational Record Form",
            "doc_no": "DOC-2026-0301-A",
            "date": "2026-03-01",
            "headers": default_headers,
            "rows": [
                ["1", "Item A (standard)", "REF-0101", "1,200", "Accepted"],
                ["2", "Item B (extended)", "REF-0204", "850", "Accepted"],
                ["3", "Item C (bundle)", "REF-0309", "2,400", "Accepted"],
            ],
            "is_discrepancy": False,
        },
        {
            "title": f"{args.company} - Exception Review Form",
            "doc_no": "DOC-2026-0305-DISCREPANCY",
            "date": "2026-03-05",
            "headers": default_headers,
            "rows": [
                ["1", "Item A (standard)", "REF-0101", "1,050", "Mismatch (-150)"],
                ["2", "Item D (superseded)", "REF-9901", "300", "Superseded reference"],
                ["3", "Item C (bundle)", "REF-0309", "2,160", "Mismatch (-240)"],
            ],
            "is_discrepancy": True,
        },
    ]
    for img_path, dflt in zip((img1_path, img2_path), scan_defaults):
        sc = scan_specs.pop(0) if scan_specs else {}
        generate_document_image(
            img_path,
            sc.get("title", dflt["title"]),
            args.company,
            sc.get("doc_no", dflt["doc_no"]),
            sc.get("date", dflt["date"]),
            sc.get("headers", dflt["headers"]),
            sc.get("rows", dflt["rows"]),
            is_discrepancy=sc.get("is_discrepancy", dflt["is_discrepancy"]),
            style=style,
        )

    # Upload all 4 files to Google Drive and grant permissions
    upload_res = upload_to_google_drive(args.company, args.suffix, [pdf_path, xlsx_path, img1_path, img2_path])

    # Save summary to JSON for skill consumption
    summary_path = out_dir / "drive_upload_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(upload_res, f, indent=2, ensure_ascii=False)
    print(f"\n💾 Upload summary saved to: {summary_path}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
